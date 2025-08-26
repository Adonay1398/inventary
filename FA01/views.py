# type: ignore
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.contrib.auth.models import User
from django.db.models import Q
from .models import Asset, Location, Movement, UserProfile, Sucursal, DispositivoSucursal, AssetImage, Responsibility
from django.utils import timezone
import csv
from django.http import HttpResponse, JsonResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from datetime import datetime
import nmap
from django.views.decorators.csrf import csrf_exempt
from .network_scanner import NetworkScanner
from django.views.decorators.http import require_POST
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .serializers import DispositivoSucursalSerializer, SucursalSerializer
from django.contrib.auth import logout
import logging
import os


logger = logging.getLogger(__name__)

@login_required
def index(request):
    """Vista principal del sistema de inventario"""
    total_assets = Asset.objects.count()
    assets_in_use = Asset.objects.filter(status='in_use').count()
    assets_in_repair = Asset.objects.filter(status='repair').count()
    recent_movements = Movement.objects.all().order_by('-movement_date')[:5]
    
    context = {
        'total_assets': total_assets,
        'assets_in_use': assets_in_use,
        'assets_in_repair': assets_in_repair,
        'recent_movements': recent_movements,
    }
    return render(request, 'FA01/index.html', context)

@login_required
def asset_list(request):
    """Lista todos los activos con opciones de filtrado"""
    assets = Asset.objects.all()
    query = request.GET.get('q')
    category = request.GET.get('category')
    status = request.GET.get('status')
    location = request.GET.get('location')
    
    if query:
        assets = assets.filter(
            Q(name__icontains=query) |
            Q(serial_number__icontains=query)
        )
    
    if category:
        assets = assets.filter(category=category)
    
    if status:
        assets = assets.filter(status=status)

    if location:
        assets = assets.filter(location_id=location)

    context = {
        'assets': assets,
        'categories': Asset.CATEGORIES,
        'status_choices': Asset.STATUS_CHOICES,
        'locations': Location.objects.all(),
    }
    return render(request, 'FA01/asset_list.html', context)

@login_required
def asset_detail(request, pk):
    """Muestra los detalles de un activo específico"""
    asset = get_object_or_404(Asset, pk=pk)
    movements = Movement.objects.filter(asset=asset).order_by('-movement_date')
    
    context = {
        'asset': asset,
        'movements': movements,
    }
    return render(request, 'FA01/asset_detail.html', context)

@login_required
def asset_create(request):
    """Crea un nuevo activo"""
    if request.method == 'POST':
        try:
            # Crear el activo
            asset = Asset(
                name=request.POST['name'],
                category=request.POST['category'],
                serial_number=request.POST['serial_number'],
                brand=request.POST.get('brand', ''),
                model=request.POST.get('model', ''),
                purchase_date=request.POST['purchase_date'],
                status=request.POST['status'],
                description=request.POST.get('description', ''),
                specifications=request.POST.get('specifications', ''),
                quantity=int(request.POST.get('quantity', 1)),
                preferred_usage_period=int(request.POST.get('preferred_usage_period', 36)),
                notes=request.POST.get('notes', '')
            )
            
            # Asignar ubicación si se proporcionó
            location_id = request.POST.get('location')
            if location_id:
                asset.location = Location.objects.get(id=location_id)
            
            # Asignar responsable si se proporcionó
            assigned_to_id = request.POST.get('assigned_to')
            if assigned_to_id:
                asset.assigned_to = User.objects.get(id=assigned_to_id)
            
            # Manejar la fecha de vencimiento de garantía si se proporcionó
            warranty_expiration = request.POST.get('warranty_expiration')
            if warranty_expiration:
                asset.warranty_expiration = warranty_expiration
            
            # Guardar el asset ANTES de crear imágenes
            asset.save()
            
            # Ahora sí puedes crear las imágenes relacionadas
            if 'images' in request.FILES:
                for image_file in request.FILES.getlist('images'):
                    try:
                        # Validar tipo de archivo
                        if not image_file.content_type.startswith('image/'):
                            messages.warning(request, f'El archivo {image_file.name} no es una imagen válida')
                            continue
                        
                        # Validar tamaño (máximo 5MB)
                        if image_file.size > 5 * 1024 * 1024:
                            messages.warning(request, f'La imagen {image_file.name} es demasiado grande (máximo 5MB)')
                            continue
                            
                        AssetImage.objects.create(asset=asset, image=image_file)
                    except Exception as img_e:
                        messages.error(request, f'Error al subir la imagen {image_file.name}: {str(img_e)}')
            
            # Subir responsivas si se proporcionaron
            if 'responsibility' in request.FILES:
                for resp_file in request.FILES.getlist('responsibility'):
                    try:
                        # Validar tipo de archivo
                        if not resp_file.content_type.startswith('image/'):
                            messages.warning(request, f'El archivo {resp_file.name} no es una imagen válida')
                            continue
                        
                        # Validar tamaño (máximo 5MB)
                        if resp_file.size > 5 * 1024 * 1024:
                            messages.warning(request, f'La responsiva {resp_file.name} es demasiado grande (máximo 5MB)')
                            continue
                            
                        Responsibility.objects.create(asset=asset, letter_responsibility=resp_file)
                    except Exception as resp_e:
                        messages.error(request, f'Error al subir la responsiva {resp_file.name}: {str(resp_e)}')
            
            # Registrar el movimiento inicial
            if location_id:
                Movement.objects.create(
                    asset=asset,
                    from_location=None,
                    to_location=asset.location,
                    from_user=None,
                    assigned_to_name=asset.assigned_to.username if asset.assigned_to else None,
                    notes='Registro inicial del activo'
                )
            
            messages.success(request, 'Activo creado exitosamente')
            return redirect('asset_detail', pk=asset.pk)
            
        except Exception as e:
            messages.error(request, f'Error al crear el activo: {str(e)}')
    
    context = {
        'categories': Asset.CATEGORIES,
        'status_choices': Asset.STATUS_CHOICES,
        'locations': Location.objects.all(),
    }
    return render(request, 'FA01/asset_form.html', context)

@login_required
def asset_update(request, pk):
    """Actualiza un activo existente"""
    asset = get_object_or_404(Asset, pk=pk)
    if request.method == 'POST':
        try:
            # Verificar si el serial_number ya existe en otro activo
            new_serial_number = request.POST.get('serial_number', '').strip()
            if new_serial_number and new_serial_number != asset.serial_number:
                existing_asset = Asset.objects.filter(serial_number=new_serial_number).exclude(pk=asset.pk).first()
                if existing_asset:
                    messages.error(request, f'El número de serie {new_serial_number} ya está en uso por otro activo')
                    raise ValueError('Serial number already exists')

            # Actualizar campos solo si vienen en POST y no están vacíos
            if request.POST.get('name'):
                asset.name = request.POST['name']
            if request.POST.get('category'):
                asset.category = request.POST['category']
            if new_serial_number:
                asset.serial_number = new_serial_number
            if request.POST.get('brand') is not None:
                asset.brand = request.POST.get('brand', asset.brand)
            if request.POST.get('model') is not None:
                asset.model = request.POST.get('model', asset.model)
            if request.POST.get('purchase_date'):
                asset.purchase_date = request.POST['purchase_date']
            if request.POST.get('status'):
                asset.status = request.POST['status']
            if request.POST.get('description') is not None:
                asset.description = request.POST.get('description', asset.description)
            if request.POST.get('specifications') is not None:
                asset.specifications = request.POST.get('specifications', asset.specifications)
            if request.POST.get('quantity'):
                asset.quantity = int(request.POST.get('quantity', asset.quantity))
            if request.POST.get('preferred_usage_period'):
                asset.preferred_usage_period = int(request.POST.get('preferred_usage_period', asset.preferred_usage_period))
            if request.POST.get('notes') is not None:
                asset.notes = request.POST.get('notes', asset.notes)

            # Actualizar ubicación
            location_id = request.POST.get('location')
            old_location = asset.location
            if location_id:
                new_location = Location.objects.get(id=location_id)
                if old_location != new_location:
                    asset.location = new_location
                    Movement.objects.create(
                        asset=asset,
                        from_location=old_location,
                        to_location=new_location,
                        from_user=asset.assigned_to,
                        assigned_to_name=asset.assigned_to.username if asset.assigned_to else None,
                        notes='Cambio de ubicación'
                    )
            else:
                asset.location = None

            # Actualizar responsable
            assigned_to_id = request.POST.get('assigned_to')
            old_assigned_to = asset.assigned_to
            if assigned_to_id:
                new_assigned_to = User.objects.get(id=assigned_to_id)
                if old_assigned_to != new_assigned_to:
                    asset.assigned_to = new_assigned_to
                    Movement.objects.create(
                        asset=asset,
                        from_location=asset.location,
                        to_location=asset.location,
                        from_user=old_assigned_to,
                        assigned_to_name=new_assigned_to.username if new_assigned_to else None,
                        notes='Cambio de responsable'
                    )
            else:
                asset.assigned_to = None

            # Actualizar responsable como texto
            assigned_to_name = request.POST.get('assigned_to_name')
            if assigned_to_name is not None:
                asset.assigned_to_name = assigned_to_name

            # Manejar nuevas imágenes
            if 'images' in request.FILES:
                for image_file in request.FILES.getlist('images'):
                    try:
                        if not image_file.content_type.startswith('image/'):
                            messages.warning(request, f'El archivo {image_file.name} no es una imagen válida')
                            continue
                        if image_file.size > 5 * 1024 * 1024:
                            messages.warning(request, f'La imagen {image_file.name} es demasiado grande (máximo 5MB)')
                            continue
                        AssetImage.objects.create(asset=asset, image=image_file)
                    except Exception as img_e:
                        messages.error(request, f'Error al subir la imagen {image_file.name}: {str(img_e)}')

            # Manejar nuevas responsivas
            if 'responsibility' in request.FILES:
                for resp_file in request.FILES.getlist('responsibility'):
                    try:
                        if not resp_file.content_type.startswith('image/'):
                            messages.warning(request, f'El archivo {resp_file.name} no es una imagen válida')
                            continue
                        if resp_file.size > 5 * 1024 * 1024:
                            messages.warning(request, f'La responsiva {resp_file.name} es demasiado grande (máximo 5MB)')
                            continue
                        Responsibility.objects.create(asset=asset, letter_responsibility=resp_file)
                    except Exception as resp_e:
                        messages.error(request, f'Error al subir la responsiva {resp_file.name}: {str(resp_e)}')

            # Actualizar fecha de vencimiento de garantía
            warranty_expiration = request.POST.get('warranty_expiration')
            if warranty_expiration:
                asset.warranty_expiration = warranty_expiration
            else:
                asset.warranty_expiration = None

            asset.save()
            messages.success(request, 'Activo actualizado exitosamente')
            return redirect('asset_detail', pk=asset.pk)
        except Exception as e:
            messages.error(request, f'Error al actualizar el activo: {str(e)}')
    context = {
        'asset': asset,
        'categories': Asset.CATEGORIES,
        'status_choices': Asset.STATUS_CHOICES,
        'locations': Location.objects.all(),
    }
    return render(request, 'FA01/asset_form.html', context)

@login_required
def location_list(request):
    """Lista todas las ubicaciones"""
    locations = Location.objects.all()
    context = {
        'locations': locations,
        'location_types': Location.LOCATION_TYPES,
    }
    return render(request, 'FA01/location_list.html', context)

@login_required
def location_create(request):
    """Crea una nueva ubicación"""
    if request.method == 'POST':
        try:
            location = Location(
                name=request.POST['name'],
                location_type=request.POST['location_type'],
                description=request.POST.get('description', '')
            )
            location.save()
            messages.success(request, 'Ubicación creada exitosamente')
        except Exception as e:
            messages.error(request, f'Error al crear la ubicación: {str(e)}')
    return redirect('location_list')

@login_required
def location_update(request, pk):
    """Actualiza una ubicación existente"""
    location = get_object_or_404(Location, pk=pk)
    if request.method == 'POST':
        try:
            location.name = request.POST['name']
            location.location_type = request.POST['location_type']
            location.description = request.POST.get('description', '')
            location.save()
            messages.success(request, 'Ubicación actualizada exitosamente')
        except Exception as e:
            messages.error(request, f'Error al actualizar la ubicación: {str(e)}')
    return redirect('location_list')

@login_required
def movement_create(request):
    """Registra un nuevo movimiento de activo"""
    if request.method == 'POST':
        try:
            asset = Asset.objects.get(id=request.POST['asset'])
            movement_type = request.POST['movement_type']
            movement = Movement(
                asset=asset,
                notes=request.POST.get('reason', ''),
                movement=movement_type,
                assigned_to_name=request.POST.get('assigned_to_name', '')
            )
            # Manejar el tipo de movimiento
            if movement_type == 'location':
                movement.from_location = asset.location
                movement.to_location = Location.objects.get(id=request.POST['to_location'])
                movement.from_user = asset.assigned_to
            elif movement_type == 'assignment':
                movement.from_location = asset.location
                movement.to_location = asset.location
                movement.from_user = asset.assigned_to
            elif movement_type == 'maintenance':
                movement.from_location = asset.location
                movement.to_location = None
                movement.from_user = asset.assigned_to
                asset.status = 'maintenance'
            elif movement_type == 'return':
                movement.from_location = None
                movement.to_location = Location.objects.get(id=request.POST['to_location'])
                movement.from_user = None
                asset.status = 'active'
            elif movement_type == 'retirement':
                movement.from_location = asset.location
                movement.to_location = None
                movement.from_user = asset.assigned_to
                asset.status = 'retired'
            # Actualizar el activo
            if movement_type == 'location':
                asset.location = movement.to_location
            elif movement_type == 'assignment':
                pass
            elif movement_type == 'return':
                asset.location = movement.to_location
            movement.save()
            asset.save()
            messages.success(request, 'Movimiento registrado exitosamente')
            return redirect('asset_detail', pk=asset.pk)
        except Exception as e:
            messages.error(request, f'Error al registrar el movimiento: {str(e)}')
    context = {
        'assets': Asset.objects.all(),
        'locations': Location.objects.all(),
        'movement_types': Movement.MOVEMENT_TYPES,
    }
    return render(request, 'FA01/movement_form.html', context)

@login_required
def user_profile(request):
    """Muestra y permite editar el perfil del usuario"""
    profile, created = UserProfile.objects.get_or_create(user=request.user)
    
    if request.method == 'POST':
        try:
            profile.department = request.POST['department']
            profile.position = request.POST['position']
            profile.extension = request.POST.get('extension', '')
            profile.save()
            
            messages.success(request, 'Perfil actualizado exitosamente')
            return redirect('user_profile')
            
        except Exception as e:
            messages.error(request, f'Error al actualizar el perfil: {str(e)}')
    
    return render(request, 'FA01/user_profile.html', {'profile': profile})

@login_required
def export_assets_excel(request):
    """Export assets to Excel file"""
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename=inventario_{datetime.now().strftime("%Y%m%d")}.xlsx'

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"

    # Define headers
    headers = [
        'ID', 'Nombre', 'Categoría', 'Número de Serie', 'Fecha de Compra',
        'Estado', 'Ubicación', 'Responsable', 'Descripción', 'Especificaciones',
        'Cantidad', 'Período de Uso Preferente', 'Vencimiento de Garantía'
    ]

    # Style for headers
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    # Write data
    assets = Asset.objects.all()
    for row, asset in enumerate(assets, 2):
        ws.cell(row=row, column=1, value=asset.id)
        ws.cell(row=row, column=2, value=asset.name)
        ws.cell(row=row, column=3, value=asset.get_category_display())
        ws.cell(row=row, column=4, value=asset.serial_number)
        ws.cell(row=row, column=5, value=asset.purchase_date.strftime('%d/%m/%Y') if asset.purchase_date else '')
        ws.cell(row=row, column=6, value=asset.get_status_display())
        ws.cell(row=row, column=7, value=asset.location.name if asset.location else '')
        ws.cell(row=row, column=8, value=asset.assigned_to.username if asset.assigned_to else '')
        ws.cell(row=row, column=9, value=asset.description)
        ws.cell(row=row, column=10, value=asset.specifications)
        ws.cell(row=row, column=11, value=asset.quantity)
        ws.cell(row=row, column=12, value=asset.preferred_usage_period)
        ws.cell(row=row, column=13, value=asset.warranty_expiration.strftime('%d/%m/%Y') if asset.warranty_expiration else '')

    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

    wb.save(response)
    return response

@login_required
def export_assets_template(request):
    """Export assets template Excel file"""
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename=plantilla_activos_{datetime.now().strftime("%Y%m%d")}.xlsx'

    wb = Workbook()
    ws = wb.active
    ws.title = "Plantilla Activos"

    # Define headers
    headers = [
        'ID', 'Nombre', 'Categoría', 'Número de Serie', 'Fecha de Compra',
        'Estado', 'Ubicación', 'Responsable', 'Descripción', 'Especificaciones',
        'Cantidad', 'Período de Uso Preferente', 'Vencimiento de Garantía'
    ]

    # Style for headers
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    # Write example data
    example_data = [
        ['', 'Laptop Dell Inspiron', 'Laptop', 'DELL123456', '15/01/2023', 'Activo', 'Oficina Principal', 'Juan Pérez', 'Laptop para desarrollo', 'Intel i5, 8GB RAM, 256GB SSD', 1, 36, '15/01/2026'],
        ['', 'Monitor HP 24"', 'Monitor', 'HP789012', '20/02/2023', 'En Uso', 'Sala de Reuniones', 'María García', 'Monitor para presentaciones', '24 pulgadas, Full HD', 1, 48, '20/02/2026'],
        ['', 'Impresora HP LaserJet', 'Impresora', 'HP345678', '10/03/2023', 'Activo', 'Recepción', 'Carlos López', 'Impresora principal', 'Láser, monocromática', 1, 60, '10/03/2026'],
    ]

    for row, data in enumerate(example_data, 2):
        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=value)

    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

    wb.save(response)
    return response

@login_required
def import_assets_excel(request):
    """Import assets from Excel file"""
    if request.method == 'POST' and request.FILES.get('excel_file'):
        try:
            excel_file = request.FILES['excel_file']
            wb = load_workbook(excel_file)
            ws = wb.active

            # Skip header row
            for row in ws.iter_rows(min_row=2):
                # Validar que tenemos datos mínimos
                if not row[1].value or not row[3].value:  # name y serial_number son obligatorios
                    continue
                
                # Procesar fecha de compra
                purchase_date = None
                if row[4].value:
                    if isinstance(row[4].value, str):
                        try:
                            # Intentar parsear diferentes formatos de fecha
                            for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y', '%Y/%m/%d']:
                                try:
                                    purchase_date = datetime.strptime(row[4].value, fmt).date()
                                    break
                                except ValueError:
                                    continue
                        except:
                            purchase_date = None
                    elif hasattr(row[4].value, 'date'):
                        purchase_date = row[4].value.date()
                
                # Procesar fecha de vencimiento de garantía
                warranty_expiration = None
                if row[12].value:  # Cambiado de 11 a 12
                    if isinstance(row[12].value, str):
                        try:
                            from datetime import datetime
                            for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y', '%Y/%m/%d']:
                                try:
                                    warranty_expiration = datetime.strptime(row[12].value, fmt).date()
                                    break
                                except ValueError:
                                    continue
                        except:
                            warranty_expiration = None
                    elif hasattr(row[12].value, 'date'):
                        warranty_expiration = row[12].value.date()
                
                # Mapear categoría
                category_mapping = {
                    'PC': 'pc',
                    'Laptop': 'laptop', 
                    'Monitor': 'monitor',
                    'Nobreak': 'nobreak',
                    'Impresora': 'printer',
                    'Equipo de Red': 'network',
                    'Periférico': 'peripheral',
                    'Servidor': 'server',
                    'Otro': 'other'
                }
                
                category = 'other'  # default
                if row[2].value:
                    category_display = str(row[2].value).strip()
                    category = category_mapping.get(category_display, 'other')
                
                # Mapear estado
                status_mapping = {
                    'Activo': 'active',
                    'En Uso': 'in_use',
                    'En Mantenimiento': 'maintenance',
                    'En Reparación': 'repair',
                    'Retirado': 'retired',
                    'Perdido': 'lost'
                }
                
                status = 'active'  # default
                if row[5].value:
                    status_display = str(row[5].value).strip()
                    status = status_mapping.get(status_display, 'active')
                
                # Procesar cantidad
                quantity = 1
                if row[10].value:
                    try:
                        quantity = int(row[10].value)
                        if quantity < 1:
                            quantity = 1
                    except (ValueError, TypeError):
                        quantity = 1
                
                # Procesar período de uso preferente
                preferred_usage_period = 36  # default
                if row[11].value:  # Nueva columna
                    try:
                        preferred_usage_period = int(row[11].value)
                        if preferred_usage_period < 1:
                            preferred_usage_period = 36
                    except (ValueError, TypeError):
                        preferred_usage_period = 36
                
                # Procesar ubicación
                location = None
                if row[6].value:  # Columna G - Ubicación
                    location_name = str(row[6].value).strip()
                    if location_name:
                        location, created = Location.objects.get_or_create(
                            name=location_name,
                            defaults={'location_type': 'office', 'description': ''}
                        )
                
                # Procesar responsable (solo como texto, no como usuario)
                assigned_to_name = None
                if row[7].value:  # Columna H - Responsable
                    assigned_to_name = str(row[7].value).strip()
                
                asset_data = {
                    'name': str(row[1].value).strip(),
                    'category': category,
                    'serial_number': str(row[3].value).strip(),
                    'purchase_date': purchase_date,
                    'status': status,
                    'description': str(row[8].value).strip() if row[8].value else '',
                    'specifications': str(row[9].value).strip() if row[9].value else '',
                    'quantity': quantity,
                    'preferred_usage_period': preferred_usage_period,
                    'warranty_expiration': warranty_expiration,
                    'location': location,
                    'assigned_to_name': assigned_to_name,
                }

                # Create or update asset
                asset, created = Asset.objects.update_or_create(
                    serial_number=asset_data['serial_number'],
                    defaults=asset_data
                )

            messages.success(request, 'Archivo Excel importado exitosamente')
        except Exception as e:
            messages.error(request, f'Error al importar el archivo: {str(e)}')

    return redirect('asset_list')

@login_required
@csrf_exempt
def network_scan(request):
    devices = []
    error = None
    if request.method == 'POST':
        network = request.POST.get('network', '192.168.1.0/24')
        try:
            nm = nmap.PortScanner()
            nm.scan(hosts=network, arguments='-sn')
            for host in nm.all_hosts():
                devices.append({
                    'ip': host,
                    'mac': nm[host]['addresses'].get('mac', ''),
                    'hostname': nm[host].hostname() if nm[host].hostname() else '',
                })
        except Exception as e:
            error = str(e)
    return render(request, 'FA01/network_scan.html', {'devices': devices, 'error': error})

def network_devices(request):
    scanner = NetworkScanner()
    devices = scanner.scan_network()
    return render(request, 'FA01/network_devices.html', {'devices': devices})

@login_required
@require_POST
def add_network_device(request):
    try:
        # Obtener datos del formulario
        ip = request.POST.get('ip')
        hostname = request.POST.get('hostname')
        device_type = request.POST.get('type')
        vendor = request.POST.get('vendor')
        model = request.POST.get('model')
        location = request.POST.get('location')
        department = request.POST.get('department')
        notes = request.POST.get('notes')

        # Crear nuevo activo en el inventario
        asset = Asset.objects.create(
            name=hostname,
            type=device_type,
            manufacturer=vendor,
            model=model,
            ip_address=ip,
            location=location,
            department=department,
            notes=notes,
            status='Activo'
        )

        return JsonResponse({
            'success': True,
            'message': 'Dispositivo agregado exitosamente'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

class RegistroDispositivosAPIView(APIView):
    def post(self, request):
        # Validar token
        auth = request.headers.get('Authorization', '')
        if not auth.startswith('Bearer '):
            return Response({'detail': 'Token no proporcionado'}, status=401)
        token = auth.split(' ')[1]
        try:
            sucursal = Sucursal.objects.get(token=token)
        except Sucursal.DoesNotExist:
            return Response({'detail': 'Token inválido o no autorizado'}, status=403)

        dispositivos = request.data.get('dispositivos', [])
        fecha_envio = request.data.get('fecha_envio', timezone.now())
        for d in dispositivos:
            DispositivoSucursal.objects.create(
                sucursal=sucursal,
                fecha_envio=fecha_envio,
                ip=d['ip'],
                mac=d['mac'],
                hostname=d['hostname']
            )
        return Response({'mensaje': 'Datos registrados exitosamente'})

class SucursalDispositivosAPIView(APIView):
    def get(self, request, codigo):
        sucursal = get_object_or_404(Sucursal, codigo=codigo)
        dispositivos = sucursal.dispositivos.all()
        serializer = DispositivoSucursalSerializer(dispositivos, many=True)
        return Response({
            'sucursal': sucursal.nombre,
            'codigo': sucursal.codigo,
            'responsable': sucursal.responsable,
            'dispositivos': serializer.data
        })

@login_required
def delete_asset_image(request, image_id):
    """Elimina una imagen específica de un activo"""
    if request.method == 'POST':
        try:
            asset_image = get_object_or_404(AssetImage, id=image_id)
            asset_id = asset_image.asset.id
            # Eliminar el archivo físico del sistema de archivos
            if asset_image.image:
                if os.path.exists(asset_image.image.path):
                    os.remove(asset_image.image.path)
            asset_image.delete()
            messages.success(request, 'Imagen eliminada exitosamente')
            return redirect('asset_update', pk=asset_id)
        except Exception as e:
            messages.error(request, f'Error al eliminar la imagen: {str(e)}')
            return redirect('asset_list')
    return redirect('asset_list')

@login_required
def delete_asset_letter_responsibility(request, image_id):
    """Elimina una responsiva específica de un activo"""
    if request.method == 'POST':
        try:
            letter_responsibility = get_object_or_404(Responsibility, id=image_id)
            asset_id = letter_responsibility.asset.id
            # Eliminar el archivo físico del sistema de archivos
            if letter_responsibility.letter_responsibility:
                if os.path.exists(letter_responsibility.letter_responsibility.path):
                    os.remove(letter_responsibility.letter_responsibility.path)
            letter_responsibility.delete()
            messages.success(request, 'Responsiva eliminada exitosamente')
            return redirect('asset_update', pk=asset_id)
        except Exception as e:
            messages.error(request, f'Error al eliminar la responsiva: {str(e)}')
            return redirect('asset_list')
    return redirect('asset_list')

@login_required
def export_locations_excel(request):
    """Export locations to Excel file"""
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename=ubicaciones_{datetime.now().strftime("%Y%m%d")}.xlsx'

    wb = Workbook()
    ws = wb.active
    ws.title = "Ubicaciones"

    # Define headers
    headers = ['Nombre', 'Tipo de Ubicación', 'Descripción']

    # Style for headers
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    # Write data
    locations = Location.objects.all()
    for row, location in enumerate(locations, 2):
        ws.cell(row=row, column=1, value=location.name)
        ws.cell(row=row, column=2, value=location.get_location_type_display())
        ws.cell(row=row, column=3, value=location.description)

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(response)
    return response

@login_required
def import_locations_excel(request):
    """Import locations from Excel file"""
    if request.method == 'POST' and request.FILES.get('excel_file'):
        try:
            excel_file = request.FILES['excel_file']
            wb = load_workbook(excel_file)
            ws = wb.active

            # Skip header row
            for row in ws.iter_rows(min_row=2):
                location_data = {
                    'name': row[0].value,
                    'location_type': row[1].value,
                    'description': row[2].value or '',
                }

                # Map location type display names to values
                location_type_mapping = {
                    'Almacén': 'warehouse',
                    'Oficina': 'office', 
                    'Farmacia': 'drugstore'
                }
                
                if location_data['location_type'] in location_type_mapping:
                    location_data['location_type'] = location_type_mapping[location_data['location_type']]
                else:
                    # Default to office if type not recognized
                    location_data['location_type'] = 'office'

                # Create or update location
                location, created = Location.objects.update_or_create(
                    name=location_data['name'],
                    defaults=location_data
                )

            messages.success(request, 'Archivo Excel de ubicaciones importado exitosamente')
        except Exception as e:
            messages.error(request, f'Error al importar el archivo: {str(e)}')

    return redirect('location_list')

@login_required
def custom_logout(request):
    """Vista personalizada para cerrar sesión"""
    logout(request)
    messages.success(request, 'Has cerrado sesión exitosamente')
    return redirect('login')
